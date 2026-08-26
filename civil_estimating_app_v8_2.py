import streamlit as st
import pandas as pd
import sqlite3
from io import BytesIO
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title='Civil Estimating Software V7', page_icon='🏗️', layout='wide', initial_sidebar_state='expanded')
APP_DIR = Path(__file__).parent
DB_FILE = APP_DIR / 'civil_estimating_library.sqlite'
DEFAULT_BNI = APP_DIR / 'construction_costs_pages_16_to_36.xlsx'
LABOR_SELL, EQUIP_SELL, MAT_TAX, MAT_SELL, HAUL_TAX, HAUL_SELL = 1.30, 1.30, 0.09, 1.09, 0.07, 1.09
WBS = {
    '02 — Existing Conditions':['Demolition','Site Clearing'],
    '31 — Earthwork':['Excavation','Backfill','Fill / Import','Grading'],
    '32 — Exterior Improvements':['Paving','Concrete','Site Improvements','Erosion Control'],
    '33 — Utilities':['Sewer','Storm Drainage','Water','Utility Structures']}
TYPES=['Labor','Equipment','Material','Trucking']
COLS=['Type','Resource','Qty','Hours','Unit','Base Rate','Tax %','Delivery','Sell Factor','Cost','Sell','Profit','Notes']

def sf(v,d=0.0):
    try: return d if pd.isna(v) else float(v)
    except: return d

def conn():
    c=sqlite3.connect(DB_FILE)
    c.execute('''CREATE TABLE IF NOT EXISTS rates(id INTEGER PRIMARY KEY, type TEXT, resource TEXT, unit TEXT, rate REAL, notes TEXT, updated TEXT, UNIQUE(type,resource))''')
    c.commit(); return c

def rate_save(typ,res,unit,rate,notes=''):
    if not str(res).strip(): return
    c=conn(); c.execute('''INSERT INTO rates(type,resource,unit,rate,notes,updated) VALUES(?,?,?,?,?,?) ON CONFLICT(type,resource) DO UPDATE SET unit=excluded.unit,rate=excluded.rate,notes=excluded.notes,updated=excluded.updated''',(typ,str(res).strip(),str(unit or ''),sf(rate),str(notes or ''),datetime.now().isoformat(timespec='seconds'))); c.commit(); c.close()

def rate_table(typ=None):
    c=conn(); q='SELECT type,resource,unit,rate,notes,updated FROM rates'; p=[]
    if typ: q+=' WHERE type=?'; p=[typ]
    q+=' ORDER BY type,resource'; d=pd.read_sql_query(q,c,params=p); c.close(); return d

def rate_lookup(typ,res):
    d=rate_table(typ)
    if d.empty: return None
    m=d[d.resource.astype(str).str.lower()==str(res).strip().lower()]
    return None if m.empty else m.iloc[0]

@st.cache_data(show_spinner=False)
def load_bni(src):
    d=pd.read_excel(src); d.columns=[str(x).strip() for x in d.columns]
    req=['CSI Code','Item Code','Description','Unit','Manhr/Unit','Page']; miss=[x for x in req if x not in d.columns]
    if miss: raise ValueError('Missing BNI columns: '+', '.join(miss))
    d=d[req].copy()
    for c in ['CSI Code','Item Code','Description','Unit']: d[c]=d[c].fillna('').astype(str).str.strip()
    d['Manhr/Unit']=pd.to_numeric(d['Manhr/Unit'],errors='coerce'); d['Page']=d['Page'].fillna('').astype(str); return d

def blank(): return pd.DataFrame(columns=COLS)

def new_task():
    return {'id':datetime.now().timestamp(),'wbs':'33 — Utilities','category':'Sewer','description':'Installation of 2,050 LF of 8" PVC Sanitary Line','quantity':2050.0,'unit':'LF','surface':'Unpaved','depth':"0'-6'",'bni_item':'','bni_description':'','bni_page':'','bni_productivity':0.0,'resources':blank()}

def calc_row(r):
    typ=str(r.get('Type','')).strip(); q=sf(r.get('Qty')); h=sf(r.get('Hours')); rate=sf(r.get('Base Rate')); tax=sf(r.get('Tax %')); delivery=sf(r.get('Delivery')); factor=sf(r.get('Sell Factor'))
    if typ=='Labor': cost=q*h*rate; factor=factor or LABOR_SELL
    elif typ=='Equipment': cost=q*h*rate; factor=factor or EQUIP_SELL
    elif typ=='Material': tax=tax or MAT_TAX; factor=factor or MAT_SELL; cost=q*rate+(q*rate*tax)+delivery
    elif typ=='Trucking': tax=tax or HAUL_TAX; factor=factor or HAUL_SELL; cost=q*rate*(1+tax)
    else: cost=0; factor=1
    sell=cost*factor; return cost,sell,sell-cost,factor,tax

def calc_resources(d):
    if d is None or d.empty: return blank()
    d=d.copy()
    for c in COLS:
        if c not in d.columns: d[c]=''
    for c in ['Type','Resource','Unit','Notes']: d[c]=d[c].fillna('').astype(str)
    for c in ['Qty','Hours','Base Rate','Tax %','Delivery','Sell Factor']: d[c]=pd.to_numeric(d[c],errors='coerce').fillna(0.0)
    out=[]
    for _,r in d.iterrows():
        typ=str(r['Type']).strip(); name=str(r['Resource']).strip()
        if name and typ in TYPES:
            lib=rate_lookup(typ,name)
            if lib is not None:
                if not str(r['Unit']).strip(): r['Unit']=lib['unit']
                if sf(r['Base Rate'])==0: r['Base Rate']=sf(lib['rate'])
            if typ=='Labor': r['Tax %']=0; r['Sell Factor']=sf(r['Sell Factor']) or LABOR_SELL
            elif typ=='Equipment': r['Tax %']=0; r['Sell Factor']=sf(r['Sell Factor']) or EQUIP_SELL
            elif typ=='Material': r['Tax %']=sf(r['Tax %']) or MAT_TAX; r['Sell Factor']=sf(r['Sell Factor']) or MAT_SELL
            elif typ=='Trucking': r['Tax %']=sf(r['Tax %']) or HAUL_TAX; r['Sell Factor']=sf(r['Sell Factor']) or HAUL_SELL
        cost,sell,profit,factor,tax=calc_row(r); r['Tax %']=tax; r['Sell Factor']=factor; r['Cost']=cost; r['Sell']=sell; r['Profit']=profit; out.append(r[COLS].to_dict())
    return pd.DataFrame(out,columns=COLS)

def calc_task(t):
    t=dict(t); t['resources']=calc_resources(t.get('resources',blank())); r=t['resources']
    vals={k:float(r.loc[r.Type==k,'Cost'].sum()) if not r.empty else 0 for k in TYPES}
    t['labor_cost']=vals['Labor']; t['equipment_cost']=vals['Equipment']; t['material_cost']=vals['Material']; t['trucking_cost']=vals['Trucking']; t['direct_cost']=sum(vals.values()); t['total_charge']=float(r['Sell'].sum()) if not r.empty else 0; t['profit']=t['total_charge']-t['direct_cost']; t['bni_hours']=sf(t['quantity'])*sf(t['bni_productivity']); return t

def save_task(t):
    for i,x in enumerate(st.session_state.tasks):
        if x['id']==t['id']: st.session_state.tasks[i]=t; return
    st.session_state.tasks.append(t)

def autosave_rates(d):
    if d is None or d.empty:return
    for _,r in d.iterrows():
        if str(r['Type']).strip() in TYPES and str(r['Resource']).strip(): rate_save(r['Type'],r['Resource'],r['Unit'],r['Base Rate'],r['Notes'])

def style(ws,fill):
    thin=Side(style='thin',color='B7B7B7')
    for c in ws[1]: c.fill=PatternFill('solid',fgColor=fill); c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center'); c.border=Border(bottom=thin)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns:
        letter=get_column_letter(col[0].column); ws.column_dimensions[letter].width=min(max(max(len(str(x.value or '')) for x in col)+2,10),45)

def export_xlsx():
    tasks=[calc_task(t) for t in st.session_state.tasks]; out=BytesIO()
    summary=[]; detail=[]; legacy=[]
    for n,t in enumerate(tasks,1):
        summary.append({'Line':n,'WBS':t['wbs'],'Category':t['category'],'Description':t['description'],'Quantity':t['quantity'],'Unit':t['unit'],'BNI MH/Unit':t['bni_productivity'],'BNI Page':t['bni_page'],'BNI Total MH':t['bni_hours'],'Labor':t['labor_cost'],'Equipment':t['equipment_cost'],'Materials':t['material_cost'],'Trucking':t['trucking_cost'],'Direct Cost':t['direct_cost'],'Total Bid Price':t['total_charge'],'Profit':t['profit']})
        detail.append({'Line':n,'WBS':t['wbs'],'Description':t['description'],'Qty':t['quantity'],'Unit':t['unit'],'BNI MH/Unit':t['bni_productivity'],'BNI Hours':t['bni_hours'],'Labor':t['labor_cost'],'Equipment':t['equipment_cost'],'Materials':t['material_cost'],'Trucking':t['trucking_cost'],'Direct Cost':t['direct_cost'],'Bid Price':t['total_charge'],'Profit':t['profit']})
        for rn,(_,r) in enumerate(t['resources'].iterrows(),1): legacy.append({'Line':n,'Resource Line':rn,'WBS':t['wbs'],'Scope':t['description'],'Type':r['Type'],'Resource':r['Resource'],'Qty':r['Qty'],'Hours':r['Hours'],'Unit':r['Unit'],'Base Rate':r['Base Rate'],'Tax %':r['Tax %'],'Delivery':r['Delivery'],'Sell Factor':r['Sell Factor'],'Cost':None,'Sell':None,'Profit':None,'Formula / Derivation':'','Notes':r['Notes']})
    with pd.ExcelWriter(out,engine='openpyxl') as w:
        pd.DataFrame(summary).to_excel(w,index=False,sheet_name='Bid Summary'); pd.DataFrame(detail).to_excel(w,index=False,sheet_name='Estimate Detail'); pd.DataFrame(legacy).to_excel(w,index=False,sheet_name='Legacy Calc')
        pd.DataFrame([['Labor Sell Factor',st.session_state.labor_sell,'Cost × factor'],['Equipment Sell Factor',st.session_state.equipment_sell,'Cost × factor'],['Material Tax',st.session_state.material_tax,'Material base × tax'],['Material Sell Factor',st.session_state.material_sell,'Cost w/ tax & delivery × factor'],['Hauling Tax',st.session_state.hauling_tax,'Hauling base × tax'],['Hauling Sell Factor',st.session_state.hauling_sell,'Cost w/ tax × factor'],['Hours / Workday',st.session_state.hours_per_day,'Duration reference']],columns=['Assumption','Value','Purpose']).to_excel(w,index=False,sheet_name='Assumptions')
    out.seek(0); wb=load_workbook(out)
    style(wb['Bid Summary'],'1F4E78'); style(wb['Estimate Detail'],'4472C4'); style(wb['Legacy Calc'],'ED7D31'); style(wb['Assumptions'],'70AD47')
    ws=wb['Legacy Calc']; H={str(c.value):c.column for c in ws[1]}
    for row in range(2,ws.max_row+1):
        typ=f'{get_column_letter(H["Type"])}{row}'; qty=f'{get_column_letter(H["Qty"])}{row}'; hrs=f'{get_column_letter(H["Hours"])}{row}'; rate=f'{get_column_letter(H["Base Rate"])}{row}'; tax=f'{get_column_letter(H["Tax %"])}{row}'; delivery=f'{get_column_letter(H["Delivery"])}{row}'; factor=f'{get_column_letter(H["Sell Factor"])}{row}'
        ws.cell(row,H['Cost'],f'=IF({typ}="Labor",{qty}*{hrs}*{rate},IF({typ}="Equipment",{qty}*{hrs}*{rate},IF({typ}="Material",({qty}*{rate})+(({qty}*{rate})*{tax})+{delivery},IF({typ}="Trucking",{qty}*{rate}*(1+{tax}),0))))')
        ws.cell(row,H['Sell'],f'={get_column_letter(H["Cost"])}{row}*{factor}'); ws.cell(row,H['Profit'],f'={get_column_letter(H["Sell"])}{row}-{get_column_letter(H["Cost"])}{row}')
        ws.cell(row,H['Formula / Derivation'],f'=IF({typ}="Labor","Qty × Hours × Base Rate",IF({typ}="Equipment","Qty × Hours × Base Rate",IF({typ}="Material","(Qty × Base Rate) + Tax + Delivery",IF({typ}="Trucking","Qty × Base Rate × (1 + Tax %)",""))))')
        for col in [H['Cost'],H['Sell'],H['Profit']]: ws.cell(row,col).number_format='$#,##0.00'
        color={'Labor':'E2F0D9','Equipment':'D9EAF7','Material':'FCE4D6','Trucking':'FFF2CC'}.get(str(ws.cell(row,H['Type']).value),'FFFFFF')
        for col in range(1,ws.max_column+1): ws.cell(row,col).fill=PatternFill('solid',fgColor=color)
    for sh in ['Bid Summary','Estimate Detail','Legacy Calc']:
        ws=wb[sh]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value,(float,int)) and any(k in str(ws.cell(1,c.column).value or '').lower() for k in ['cost','price','sell','profit','rate','delivery']): c.number_format='$#,##0.00'
    final=BytesIO(); wb.save(final); final.seek(0); return final.getvalue()

# UI
st.markdown('''<style>#MainMenu,header,footer{visibility:hidden}.block-container{padding-top:.4rem;max-width:1800px}.ribbon{background:#1f4e78;color:white;padding:10px 14px;border-radius:6px;margin-bottom:8px}.title{font-size:21px;font-weight:700}.sub{font-size:11px}.section{background:#eaf2f8;border-left:4px solid #1f4e78;padding:7px 10px;margin:8px 0;font-weight:700}.audit{background:#fff8e1;border:1px solid #ead38b;border-radius:6px;padding:10px}.kpi{border:1px solid #d9e2f3;border-radius:6px;padding:7px}.kpi-label{font-size:10px;color:#666}.kpi-value{font-size:18px;font-weight:700}[data-testid="stSidebar"]{min-width:290px;max-width:330px}''',unsafe_allow_html=True)
st.markdown('<div class="ribbon"><div class="title">🏗️ Civil Estimating Software — Version 7</div><div class="sub">Sigma / Excel-style Unit Price Analysis • BNI Productivity • Resource Cost Breakdown • Bid Workbook</div></div>',unsafe_allow_html=True)

if 'tasks' not in st.session_state: st.session_state.tasks=[]
if 'active' not in st.session_state: st.session_state.active=None
if 'bni' not in st.session_state:
    try: st.session_state.bni=load_bni(DEFAULT_BNI) if DEFAULT_BNI.exists() else None
    except: st.session_state.bni=None
for k,v in {'project':'Civil Construction Estimate','estimator':'','labor_sell':LABOR_SELL,'equipment_sell':EQUIP_SELL,'material_tax':MAT_TAX,'material_sell':MAT_SELL,'hauling_tax':HAUL_TAX,'hauling_sell':HAUL_SELL,'hours_per_day':8.0}.items():
    if k not in st.session_state: st.session_state[k]=v

with st.sidebar:
    st.markdown('## 📂 Project'); st.session_state.project=st.text_input('Project',st.session_state.project); st.session_state.estimator=st.text_input('Estimator',st.session_state.estimator)
    if st.button('➕ NEW SCOPE ITEM',type='primary',use_container_width=True):
        t=new_task(); st.session_state.tasks.append(t); st.session_state.active=t['id']; st.rerun()
    st.divider(); st.markdown('### WBS / Scope Tree')
    for master,children in WBS.items():
        with st.expander(master,expanded=True):
            for child in children:
                st.caption('▸ '+child)
                for t in st.session_state.tasks:
                    if t['wbs']==master and t['category']==child:
                        label=t['description'][:46]+'...' if len(t['description'])>49 else t['description']
                        if st.button(label,key=f'tree_{t["id"]}',use_container_width=True): st.session_state.active=t['id']; st.rerun()

ribbon=st.tabs(['Estimate Sheet','Bid Summary','Cost Libraries / BNi','Assumptions','Excel Export'])

def current():
    for t in st.session_state.tasks:
        if t['id']==st.session_state.active:return t
    return None

with ribbon[0]:
    t=current()
    if t is None: st.info('Click **NEW SCOPE ITEM** in the sidebar to begin.')
    else:
        t=calc_task(t); k=st.columns(7)
        for c,(lab,val) in zip(k,[('BNI Hours',f'{t["bni_hours"]:,.2f} MH'),('Labor',f'${t["labor_cost"]:,.2f}'),('Materials',f'${t["material_cost"]:,.2f}'),('Equipment',f'${t["equipment_cost"]:,.2f}'),('Trucking',f'${t["trucking_cost"]:,.2f}'),('Direct Cost',f'${t["direct_cost"]:,.2f}'),('Profit',f'${t["profit"]:,.2f}')]): c.metric(lab,val)
        st.success(f'TOTAL BID PRICE / CHARGE: ${t["total_charge"]:,.2f}')
        st.markdown('<div class="section">1. SCOPE HEADER</div>',unsafe_allow_html=True)
        c1,c2=st.columns([2,1]); t['description']=c1.text_input('Scope Description',t['description']); cats=[x for v in WBS.values() for x in v]; t['category']=c2.selectbox('Category',cats,index=cats.index(t['category']) if t['category'] in cats else 0)
        c1,c2,c3,c4=st.columns(4); t['quantity']=c1.number_input('Scope Quantity',min_value=0.,value=sf(t['quantity']),step=1.); units=['LF','SF','SY','CY','TON','EA','HR','DAY','LS']; t['unit']=c2.selectbox('Unit',units,index=units.index(t['unit']) if t['unit'] in units else 0); t['surface']=c3.selectbox('Surface',['Paved','Unpaved','Repaired']); t['depth']=c4.text_input('Average Depth',t['depth'])
        masters=list(WBS); t['wbs']=st.selectbox('MasterFormat / WBS',masters,index=masters.index(t['wbs']) if t['wbs'] in masters else 3); st.caption('Depth is descriptive only in V7; no automatic excavation-depth calculator is applied.')
        st.markdown('<div class="section">2. BNI PRODUCTIVITY</div>',unsafe_allow_html=True)
        if st.session_state.bni is None: st.warning('Upload the BNI Excel database under Cost Libraries / BNi.')
        else:
            b=st.session_state.bni; q=st.text_input('Search BNI',placeholder='sewer / PVC / excavation / asphalt',key=f'bq_{t["id"]}'); s=b.copy()
            if q.strip():
                q=q.lower().strip(); s=s[s.Description.str.lower().str.contains(q,na=False)|s['CSI Code'].str.lower().str.contains(q,na=False)|s['Item Code'].str.lower().str.contains(q,na=False)|s.Unit.str.lower().str.contains(q,na=False)]
            s=s[s['Manhr/Unit'].notna()].head(200)
            if not s.empty:
                opts={i:f'{x.Description} | {x.Unit} | {x["Manhr/Unit"]:.4f} MH/{x.Unit} | Item {x["Item Code"]} | Page {x.Page}' for i,x in s.iterrows()}; ids=list(opts); sel=st.selectbox('Select BNI Item',ids,format_func=lambda x:opts[x],key=f'bsel_{t["id"]}'); br=b.loc[sel]; t['bni_item']=str(br['Item Code']); t['bni_description']=str(br['Description']); t['bni_page']=str(br['Page']); t['bni_productivity']=sf(br['Manhr/Unit']); st.markdown(f'<div class="audit"><b>BNI:</b> {t["bni_description"]} | Item {t["bni_item"]} | Page {t["bni_page"]}<br><b>Derivation:</b> {t["quantity"]:,.2f} {t["unit"]} × {t["bni_productivity"]:.4f} MH/{t["unit"]} = <b>{t["bni_hours"]:,.2f} MH</b></div>',unsafe_allow_html=True)
        st.markdown('<div class="section">3. RESOURCE BREAKDOWN / UNIT PRICE ANALYSIS</div>',unsafe_allow_html=True)
        res=t['resources'].copy()
        if res.empty: res=pd.DataFrame([{'Type':'Labor','Resource':'Laborer','Qty':1.,'Hours':0.,'Unit':'HR','Base Rate':0.,'Tax %':0.,'Delivery':0.,'Sell Factor':LABOR_SELL,'Cost':0.,'Sell':0.,'Profit':0.,'Notes':''}],columns=COLS)
        edited=st.data_editor(res,num_rows='dynamic',use_container_width=True,hide_index=True,key=f'ed_{t["id"]}',column_config={'Type':st.column_config.SelectboxColumn(options=TYPES),'Qty':st.column_config.NumberColumn(format='%.3f'),'Hours':st.column_config.NumberColumn(format='%.3f'),'Base Rate':st.column_config.NumberColumn(format='$%.2f'),'Tax %':st.column_config.NumberColumn(format='%.1%%'),'Delivery':st.column_config.NumberColumn(format='$%.2f'),'Sell Factor':st.column_config.NumberColumn(format='%.2fx'),'Cost':st.column_config.NumberColumn(disabled=True,format='$%.2f'),'Sell':st.column_config.NumberColumn(disabled=True,format='$%.2f'),'Profit':st.column_config.NumberColumn(disabled=True,format='$%.2f')})
        t['resources']=calc_resources(edited); autosave_rates(t['resources']); t=calc_task(t)
        st.dataframe(t['resources'],use_container_width=True,hide_index=True)
        st.markdown('<div class="section">4. DERIVATION / AUDIT TRAIL</div>',unsafe_allow_html=True)
        st.markdown(f'<div class="audit"><b>Direct Cost</b> = Labor + Equipment + Materials + Trucking = ${t["direct_cost"]:,.2f}<br><b>Total Bid Price</b> = Resource sell prices = ${t["total_charge"]:,.2f}<br><b>Estimated Profit</b> = Bid Price − Direct Cost = ${t["profit"]:,.2f}</div>',unsafe_allow_html=True)
        resource_hours=sum(sf(r.Qty)*sf(r.Hours) for _,r in t['resources'].iterrows() if r.Type in ['Labor','Equipment'])
        st.markdown('<div class="section">5. PRODUCTION / DURATION</div>',unsafe_allow_html=True); d1,d2,d3=st.columns(3); d1.metric('BNI Man-Hours',f'{t["bni_hours"]:,.2f}'); d2.metric('Resource Hours',f'{resource_hours:,.2f}'); d3.metric('Estimated Days',f'{resource_hours/st.session_state.hours_per_day:,.2f}')
        a,b,c=st.columns(3)
        if a.button('💾 SAVE SCOPE ITEM',type='primary',use_container_width=True): save_task(t); st.success('Saved.')
        if b.button('📋 DUPLICATE',use_container_width=True):
            cp=dict(t); cp['id']=datetime.now().timestamp(); cp['description']+=' — Copy'; cp['resources']=t['resources'].copy(); st.session_state.tasks.append(cp); st.session_state.active=cp['id']; st.rerun()
        if c.button('🗑️ DELETE',use_container_width=True): st.session_state.tasks=[x for x in st.session_state.tasks if x['id']!=t['id']]; st.session_state.active=st.session_state.tasks[0]['id'] if st.session_state.tasks else None; st.rerun()

with ribbon[1]:
    st.header('Bid Summary')
    if st.session_state.tasks:
        rows=[]
        for i,x in enumerate(st.session_state.tasks,1):
            t=calc_task(x); rows.append({'Line':i,'WBS':t['wbs'],'Category':t['category'],'Description':t['description'],'Qty':t['quantity'],'Unit':t['unit'],'BNI Hours':t['bni_hours'],'Labor':t['labor_cost'],'Equipment':t['equipment_cost'],'Materials':t['material_cost'],'Trucking':t['trucking_cost'],'Direct Cost':t['direct_cost'],'Total Bid Price':t['total_charge'],'Profit':t['profit']})
        d=pd.DataFrame(rows); c=st.columns(5); c[0].metric('Direct Cost',f'${d["Direct Cost"].sum():,.2f}'); c[1].metric('Total Bid Price',f'${d["Total Bid Price"].sum():,.2f}'); c[2].metric('Profit',f'${d["Profit"].sum():,.2f}'); c[3].metric('BNI Hours',f'{d["BNI Hours"].sum():,.2f}'); c[4].metric('Scope Items',len(d)); st.dataframe(d,use_container_width=True,hide_index=True); st.subheader('Category Breakdown'); st.dataframe(d.groupby('Category',as_index=False)[['Labor','Equipment','Materials','Trucking','Direct Cost','Total Bid Price','Profit']].sum(),use_container_width=True,hide_index=True)
    else: st.info('No scope items yet.')

with ribbon[2]:
    st.header('Cost Libraries / BNi'); tabs=st.tabs(['BNI Productivity','Labor','Equipment','Materials','Trucking','SQLite Library'])
    with tabs[0]:
        up=st.file_uploader('Upload / replace BNI productivity Excel',type=['xlsx']);
        if up:
            try: st.session_state.bni=load_bni(up); st.success(f'Loaded {len(st.session_state.bni):,} BNI records.')
            except Exception as e: st.error(str(e))
        if st.session_state.bni is not None: st.dataframe(st.session_state.bni.head(500),use_container_width=True,hide_index=True)
    def lib_editor(typ):
        d=rate_table(typ); 
        if d.empty: d=pd.DataFrame(columns=['type','resource','unit','rate','notes','updated'])
        e=st.data_editor(d,num_rows='dynamic',use_container_width=True,hide_index=True,key='lib_'+typ)
        if st.button('💾 SAVE '+typ.upper()+' LIBRARY',key='save_'+typ,use_container_width=True):
            for _,r in e.iterrows():
                if str(r.get('resource','')).strip(): rate_save(typ,r['resource'],r.get('unit',''),r.get('rate',0),r.get('notes',''))
            st.success('Library saved.'); st.rerun()
    with tabs[1]: lib_editor('Labor')
    with tabs[2]: lib_editor('Equipment')
    with tabs[3]: lib_editor('Material')
    with tabs[4]: lib_editor('Trucking')
    with tabs[5]: st.dataframe(rate_table(),use_container_width=True,hide_index=True)

with ribbon[3]:
    st.header('Assumptions / Legacy Calc Formula Map')
    c1, c2 = st.columns(2)

st.session_state.labor_sell = c1.number_input(
    "Labor Sell Factor",
    min_value=1.0,
    value=float(st.session_state.labor_sell),
    step=0.01,
    format="%.2f"
)

st.session_state.equipment_sell = c2.number_input(
    "Equipment Sell Factor",
    min_value=1.0,
    value=float(st.session_state.equipment_sell),
    step=0.01,
    format="%.2f"
)

st.session_state.material_tax = c1.number_input(
    "Material Tax (%)",
    min_value=0.0,
    max_value=1.0,
    value=float(st.session_state.material_tax),
    step=0.01,
    format="%.2f"
)

st.session_state.material_sell = c2.number_input(
    "Material Sell Factor",
    min_value=1.0,
    value=float(st.session_state.material_sell),
    step=0.01,
    format="%.2f"
)

st.session_state.hauling_tax = c1.number_input(
    "Hauling Tax (%)",
    min_value=0.0,
    max_value=1.0,
    value=float(st.session_state.hauling_tax),
    step=0.01,
    format="%.2f"
)

st.session_state.hauling_sell = c2.number_input(
    "Hauling Sell Factor",
    min_value=1.0,
    value=float(st.session_state.hauling_sell),
    step=0.01,
    format="%.2f"
)

st.session_state.hours_per_day = st.number_input(
    "Hours / Workday",
    min_value=1.0,
    value=float(st.session_state.hours_per_day),
    step=0.5,
    format="%.1f"
)
    st.dataframe(pd.DataFrame([['Labor','Qty × Hours × Base Rate','Cost × 1.30'],['Equipment','Qty × Hours × Base Rate','Cost × 1.30'],['Material','(Qty × Rate) + Tax + Delivery','Cost × 1.09'],['Trucking','Qty × Rate × 1.07','Cost × 1.09'],['BNI','Scope Qty × BNI Manhr/Unit','Reference hours']],columns=['Type','Cost Formula','Sell / Result']),use_container_width=True,hide_index=True)

with ribbon[4]:
    st.header('Excel Export');
    if not st.session_state.tasks: st.info('Create at least one scope item first.')
    else:
        st.write('Workbook contains Bid Summary, Estimate Detail, Legacy Calc, and Assumptions. Legacy Calc contains active Excel formulas and color-coded resource rows.')
        st.download_button('📊 DOWNLOAD VERSION 7 EXCEL ESTIMATE',export_xlsx(),file_name=f'Civil_Estimate_V7_{datetime.now():%Y%m%d_%H%M}.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',type='primary',use_container_width=True)
        st.metric('Total Bid Price',f'${sum(calc_task(x)["total_charge"] for x in st.session_state.tasks):,.2f}')

st.caption('Civil Estimating Software V7 • BNI Productivity + Resource Unit Price Analysis + Legacy Calc + Excel Export')
