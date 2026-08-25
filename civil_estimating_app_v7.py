import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from datetime import datetime
import json

# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Civil Estimating Software",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# CONSTANTS
# ============================================================

APP_VERSION = "Version 8 — Desktop Estimating Workbook"

MARKUP_MULTIPLIER = 1.30
MATERIAL_TAX = 0.09
HAULING_TAX = 0.07

DB = Path(__file__).parent / "construction_costs_pages_16_to_36.xlsx"


# ============================================================
# DESKTOP-STYLE CSS
# ============================================================

st.markdown("""
<style>

html, body, [class*="css"] {
    font-family: "Segoe UI", Arial, sans-serif;
}

/* Hide Streamlit chrome */
#MainMenu {
    visibility: hidden;
}

footer {
    visibility: hidden;
}

header {
    visibility: hidden;
}

/* Main application */
.block-container {
    padding-top: 0.8rem;
    padding-bottom: 2rem;
    max-width: 100%;
}

/* Ribbon */
.ribbon {
    background: #f3f3f3;
    border: 1px solid #c8c8c8;
    border-radius: 4px;
    padding: 4px 8px;
    margin-bottom: 8px;
    box-shadow: 0 1px 2px rgba(0,0,0,.08);
}

.ribbon-title {
    font-size: 12px;
    font-weight: 600;
    color: #555;
    padding: 2px 6px;
}

.ribbon-subtitle {
    font-size: 10px;
    color: #777;
}

/* Application title */
.app-title {
    font-size: 22px;
    font-weight: 600;
    color: #1f1f1f;
}

.app-subtitle {
    font-size: 11px;
    color: #666;
}

/* KPI cards */
.kpi {
    background: white;
    border: 1px solid #d7d7d7;
    border-radius: 4px;
    padding: 10px 12px;
    min-height: 82px;
    box-shadow: 0 1px 2px rgba(0,0,0,.04);
}

.kpi-label {
    font-size: 11px;
    color: #666;
    text-transform: uppercase;
}

.kpi-value {
    font-size: 20px;
    font-weight: 650;
    margin-top: 4px;
}

.kpi-small {
    font-size: 12px;
    color: #555;
}

/* Section headers */
.section-header {
    background: #eaf2f8;
    border: 1px solid #c8d9e6;
    padding: 7px 10px;
    margin-top: 8px;
    margin-bottom: 6px;
    font-weight: 600;
    font-size: 13px;
}

/* Status */
.status-bar {
    background: #f7f7f7;
    border-top: 1px solid #d0d0d0;
    border-bottom: 1px solid #d0d0d0;
    padding: 4px 8px;
    font-size: 11px;
    color: #666;
}

/* WBS */
.wbs-header {
    font-size: 13px;
    font-weight: 650;
    padding-bottom: 5px;
    border-bottom: 1px solid #ddd;
}

/* Resource group */
.resource-labor {
    border-left: 4px solid #4472C4;
}

.resource-material {
    border-left: 4px solid #70AD47;
}

.resource-equipment {
    border-left: 4px solid #ED7D31;
}

.resource-hauling {
    border-left: 4px solid #A5A5A5;
}

/* Estimate total */
.final-total {
    background: #e2f0d9;
    border: 1px solid #a9d18e;
    border-radius: 4px;
    padding: 14px;
    text-align: right;
}

.final-total-label {
    font-size: 12px;
    color: #555;
}

.final-total-value {
    font-size: 27px;
    font-weight: 700;
}

/* Tabs */
button[data-baseweb="tab"] {
    font-family: "Segoe UI", Arial, sans-serif;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    border-right: 1px solid #cfcfcf;
}

</style>
""", unsafe_allow_html=True)


# ============================================================
# DATABASE
# ============================================================

@st.cache_data
def load_bni(source):

    df = pd.read_excel(source)

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    required = [
        "CSI Code",
        "Item Code",
        "Description",
        "Unit",
        "Manhr/Unit",
        "Page"
    ]

    missing = [
        c for c in required
        if c not in df.columns
    ]

    if missing:
        raise ValueError(
            "Missing columns: "
            + ", ".join(missing)
        )

    df = df[required].copy()

    for c in [
        "CSI Code",
        "Item Code",
        "Description",
        "Unit"
    ]:
        df[c] = (
            df[c]
            .fillna("")
            .astype(str)
            .str.strip()
        )

    df["Manhr/Unit"] = pd.to_numeric(
        df["Manhr/Unit"],
        errors="coerce"
    )

    return df


# ============================================================
# SESSION STATE
# ============================================================

def initialize_state():

    defaults = {

        "active_task": None,

        "tasks": [],

        "active_resource_rows": pd.DataFrame(
            columns=[
                "Type",
                "Resource",
                "Quantity",
                "Unit",
                "Rate",
                "Tax %",
                "Cost",
                "Charge",
                "Profit"
            ]
        ),

        "project_name":
            "My Civil Construction Project",

        "estimator":
            "",

        "workday_hours":
            8.0,

        "markup":
            MARKUP_MULTIPLIER,

        "material_tax":
            MATERIAL_TAX,

        "hauling_tax":
            HAULING_TAX
    }

    for key, value in defaults.items():

        if key not in st.session_state:

            st.session_state[key] = value


initialize_state()


# ============================================================
# RESOURCE CALCULATION
# ============================================================

def calculate_resources(df):

    if df is None or df.empty:

        return df

    df = df.copy()

    numeric_columns = [
        "Quantity",
        "Rate",
        "Tax %",
        "Cost",
        "Charge",
        "Profit"
    ]

    for column in numeric_columns:

        if column in df.columns:

            df[column] = pd.to_numeric(
                df[column],
                errors="coerce"
            ).fillna(0.0)

    costs = []

    charges = []

    profits = []

    for _, row in df.iterrows():

        resource_type = str(
            row.get("Type", "")
        ).strip().lower()

        quantity = float(
            row.get("Quantity", 0)
        )

        rate = float(
            row.get("Rate", 0)
        )

        tax = float(
            row.get("Tax %", 0)
        )

        base_cost = quantity * rate

        tax_amount = (
            base_cost *
            tax
        )

        total_cost = (
            base_cost +
            tax_amount
        )

        charge = (
            total_cost *
            st.session_state.markup
        )

        profit = (
            charge -
            total_cost
        )

        costs.append(total_cost)
        charges.append(charge)
        profits.append(profit)

    df["Cost"] = costs
    df["Charge"] = charges
    df["Profit"] = profits

    return df


# ============================================================
# TASK CALCULATION
# ============================================================

def calculate_task(task):

    resources = task.get(
        "resources",
        pd.DataFrame()
    )

    resources = calculate_resources(
        resources
    )

    task["resources"] = resources

    labor_cost = resources.loc[
        resources["Type"].str.lower() == "labor",
        "Cost"
    ].sum()

    material_cost = resources.loc[
        resources["Type"].str.lower() == "material",
        "Cost"
    ].sum()

    equipment_cost = resources.loc[
        resources["Type"].str.lower() == "equipment",
        "Cost"
    ].sum()

    hauling_cost = resources.loc[
        resources["Type"].str.lower() == "trucking",
        "Cost"
    ].sum()

    direct_cost = (
        labor_cost
        + material_cost
        + equipment_cost
        + hauling_cost
    )

    total_charge = resources["Charge"].sum()

    total_profit = (
        total_charge -
        direct_cost
    )

    total_hours = 0.0

    if not resources.empty:

        labor_hours = resources.loc[
            resources["Type"].str.lower() == "labor",
            "Quantity"
        ].sum()

        equipment_hours = resources.loc[
            resources["Type"].str.lower() == "equipment",
            "Quantity"
        ].sum()

        hauling_hours = resources.loc[
            resources["Type"].str.lower() == "trucking",
            "Quantity"
        ].sum()

        total_hours = (
            labor_hours
            + equipment_hours
            + hauling_hours
        )

    task["labor_cost"] = labor_cost
    task["material_cost"] = material_cost
    task["equipment_cost"] = equipment_cost
    task["hauling_cost"] = hauling_cost
    task["direct_cost"] = direct_cost
    task["total_charge"] = total_charge
    task["profit"] = total_profit
    task["total_hours"] = total_hours

    return task


# ============================================================
# RIBBON
# ============================================================

st.markdown(
    """
    <div class="ribbon">
        <div class="app-title">
            🏗️ Civil Estimating Software
        </div>
        <div class="app-subtitle">
            Desktop Estimating Workbook • BNI Productivity •
            Resource Cost Assembly
        </div>
    </div>
    """,
    unsafe_allow_html=True
)


ribbon_tabs = st.tabs([
    "📁 File",
    "📚 Libraries / Price Lists",
    "📋 Estimate Sheet",
    "📊 Summary / Reports",
    "📤 Excel Export"
])


# ============================================================
# SIDEBAR — PROJECT TREE
# ============================================================

with st.sidebar:

    st.markdown(
        '<div class="wbs-header">PROJECT EXPLORER</div>',
        unsafe_allow_html=True
    )

    st.text_input(
        "Project",
        key="project_name"
    )

    st.text_input(
        "Estimator",
        key="estimator"
    )

    st.divider()

    st.markdown(
        "**MASTERFORMAT WBS**"
    )

    wbs = {
        "02 — Existing Conditions": [],
        "03 — Concrete": [],
        "31 — Earthwork": [],
        "32 — Exterior Improvements": [],
        "33 — Utilities": []
    }

    for task in st.session_state.tasks:

        category = task.get(
            "wbs",
            "33 — Utilities"
        )

        if category not in wbs:

            wbs[category] = []

        wbs[category].append(task)


    for category, category_tasks in wbs.items():

        with st.expander(
            category,
            expanded=True
        ):

            if not category_tasks:

                st.caption(
                    "No estimate items"
                )

            for index, task in enumerate(
                category_tasks
            ):

                label = task.get(
                    "description",
                    "Unnamed Task"
                )

                if st.button(
                    label[:48],
                    key=f"task_{category}_{index}",
                    use_container_width=True
                ):

                    st.session_state.active_task = task


    st.divider()

    if st.button(
        "➕ NEW ESTIMATE ITEM",
        use_container_width=True
    ):

        st.session_state.active_task = {
            "id":
                datetime.now().timestamp(),

            "wbs":
                "33 — Utilities",

            "description":
                "New Estimate Item",

            "quantity":
                1.0,

            "unit":
                "EA",

            "bni_productivity":
                0.0,

            "bni_page":
                "",

            "resources":
                pd.DataFrame(
                    columns=[
                        "Type",
                        "Resource",
                        "Quantity",
                        "Unit",
                        "Rate",
                        "Tax %",
                        "Cost",
                        "Charge",
                        "Profit"
                    ]
                )
        }


# ============================================================
# FILE TAB
# ============================================================

with ribbon_tabs[0]:

    st.subheader("File")

    col1, col2, col3 = st.columns(3)

    with col1:

        st.write("### New Project")

        if st.button(
            "Create New Project",
            use_container_width=True
        ):

            st.session_state.tasks = []

            st.session_state.active_task = None

            st.success(
                "New project created."
            )

    with col2:

        st.write("### Project Information")

        st.write(
            f"**Project:** "
            f"{st.session_state.project_name}"
        )

        st.write(
            f"**Estimator:** "
            f"{st.session_state.estimator}"
        )

    with col3:

        st.write("### Version")

        st.info(
            APP_VERSION
        )


# ============================================================
# LIBRARIES TAB
# ============================================================

with ribbon_tabs[1]:

    st.subheader(
        "📚 Libraries / Price Lists"
    )

    lib_tabs = st.tabs([
        "BNI Productivity",
        "Labor",
        "Materials",
        "Equipment",
        "Trucking"
    ])

    # --------------------------------------------------------
    # BNI
    # --------------------------------------------------------

    with lib_tabs[0]:

        uploaded = st.file_uploader(
            "BNI Productivity Database",
            type=["xlsx"],
            key="bni_upload"
        )

        try:

            if uploaded:

                bni_df = load_bni(
                    uploaded
                )

            elif DB.exists():

                bni_df = load_bni(DB)

            else:

                bni_df = pd.DataFrame()

        except Exception as e:

            st.error(str(e))

            bni_df = pd.DataFrame()


        if not bni_df.empty:

            st.success(
                f"{len(bni_df):,} BNI records loaded."
            )

            st.dataframe(
                bni_df.head(100),
                use_container_width=True,
                hide_index=True
            )


    # --------------------------------------------------------
    # Labor
    # --------------------------------------------------------

    with lib_tabs[1]:

        labor_rates = st.data_editor(

            pd.DataFrame([
                ["Foreman", "HR", 55.00],
                ["Laborer", "HR", 42.00],
                ["Equipment Operator", "HR", 48.00],
            ], columns=[
                "Resource",
                "Unit",
                "Rate"
            ]),

            num_rows="dynamic",
            use_container_width=True
        )


    # --------------------------------------------------------
    # Materials
    # --------------------------------------------------------

    with lib_tabs[2]:

        material_rates = st.data_editor(

            pd.DataFrame([
                ['8" PVC', "LF", 0.00],
                ["8\" Sewer Tap", "EA", 0.00],
                ["Crushed Gravel", "CY", 32.00],
                ["Import Fill", "CY", 0.00],
            ], columns=[
                "Resource",
                "Unit",
                "Rate"
            ]),

            num_rows="dynamic",
            use_container_width=True
        )


    # --------------------------------------------------------
    # Equipment
    # --------------------------------------------------------

    with lib_tabs[3]:

        equipment_rates = st.data_editor(

            pd.DataFrame([
                ["307.5", "HR", 125.00],
                ["CTL", "HR", 110.00],
                ["Excavator", "HR", 125.00],
                ["Loader", "HR", 110.00],
            ], columns=[
                "Resource",
                "Unit",
                "Rate"
            ]),

            num_rows="dynamic",
            use_container_width=True
        )


    # --------------------------------------------------------
    # Trucking
    # --------------------------------------------------------

    with lib_tabs[4]:

        trucking_rates = st.data_editor(

            pd.DataFrame([
                ["Tri-Axle", "HR", 95.00],
                ["Dump Truck", "HR", 95.00],
            ], columns=[
                "Resource",
                "Unit",
                "Rate"
            ]),

            num_rows="dynamic",
            use_container_width=True
        )


# ============================================================
# ESTIMATE SHEET
# ============================================================

with ribbon_tabs[2]:

    st.subheader(
        "📋 Estimate Sheet"
    )

    task = st.session_state.active_task

    if task is None:

        st.info(
            "Select an estimate item from the "
            "Project Explorer or create a new item."
        )

    else:

        # ----------------------------------------------------
        # TASK HEADER
        # ----------------------------------------------------

        description = st.text_input(
            "Description",
            value=task.get(
                "description",
                ""
            )
        )

        quantity = st.number_input(
            "Quantity",
            min_value=0.0,
            value=float(
                task.get(
                    "quantity",
                    1.0
                )
            ),
            step=1.0
        )

        unit = st.text_input(
            "Unit",
            value=task.get(
                "unit",
                "EA"
            )
        )

        wbs = st.selectbox(
            "WBS",
            [
                "02 — Existing Conditions",
                "03 — Concrete",
                "31 — Earthwork",
                "32 — Exterior Improvements",
                "33 — Utilities"
            ],
            index=[
                "02 — Existing Conditions",
                "03 — Concrete",
                "31 — Earthwork",
                "32 — Exterior Improvements",
                "33 — Utilities"
            ].index(
                task.get(
                    "wbs",
                    "33 — Utilities"
                )
            )
        )


        # ----------------------------------------------------
        # BNI PRODUCTIVITY
        # ----------------------------------------------------

        st.markdown(
            '<div class="section-header">'
            'BNI PRODUCTIVITY REFERENCE'
            '</div>',
            unsafe_allow_html=True
        )

        bni_search = st.text_input(
            "Search BNI",
            placeholder=(
                "Example: sanitary, PVC, excavation"
            )
        )

        selected_bni = None

        if "bni_df" in locals() and not bni_df.empty:

            search_df = bni_df.copy()

            if bni_search.strip():

                q = bni_search.lower()

                search_df = search_df[
                    search_df["Description"]
                    .str.lower()
                    .str.contains(
                        q,
                        na=False
                    )
                ]

            search_df = search_df[
                search_df["Manhr/Unit"].notna()
            ].head(100)

            if not search_df.empty:

                bni_options = {

                    idx:
                        f'{row["Description"]} | '
                        f'{row["Unit"]} | '
                        f'{row["Manhr/Unit"]:.4f} MH/{row["Unit"]} | '
                        f'Page {row["Page"]}'

                    for idx, row
                    in search_df.iterrows()
                }

                selected_bni = st.selectbox(
                    "BNI Item",
                    list(
                        bni_options.keys()
                    ),
                    format_func=lambda x:
                        bni_options[x]
                )


        if selected_bni is not None:

            bni_row = bni_df.loc[
                selected_bni
            ]

            task["bni_productivity"] = float(
                bni_row["Manhr/Unit"]
            )

            task["bni_page"] = str(
                bni_row["Page"]
            )

            task["bni_description"] = str(
                bni_row["Description"]
            )

            st.info(
                f'BNI: {bni_row["Description"]}  | '
                f'{bni_row["Manhr/Unit"]:.4f} MH/{bni_row["Unit"]}  | '
                f'Page {bni_row["Page"]}'
            )


        # ----------------------------------------------------
        # RESOURCE GRID
        # ----------------------------------------------------

        st.markdown(
            '<div class="section-header">'
            'RESOURCE BREAKDOWN — DIRECT ENTRY'
            '</div>',
            unsafe_allow_html=True
        )

        resources = task.get(
            "resources",
            pd.DataFrame(
                columns=[
                    "Type",
                    "Resource",
                    "Quantity",
                    "Unit",
                    "Rate",
                    "Tax %",
                    "Cost",
                    "Charge",
                    "Profit"
                ]
            )
        )

        edited_resources = st.data_editor(

            resources,

            num_rows="dynamic",

            use_container_width=True,

            hide_index=True,

            column_config={

                "Type":
                    st.column_config.SelectboxColumn(
                        "Type",
                        options=[
                            "Labor",
                            "Material",
                            "Equipment",
                            "Trucking"
                        ]
                    ),

                "Quantity":
                    st.column_config.NumberColumn(
                        "Qty",
                        format="%.3f"
                    ),

                "Rate":
                    st.column_config.NumberColumn(
                        "Rate",
                        format="$%.2f"
                    ),

                "Tax %":
                    st.column_config.NumberColumn(
                        "Tax %",
                        format="%.1f%%"
                    ),

                "Cost":
                    st.column_config.NumberColumn(
                        "Cost",
                        format="$%.2f",
                        disabled=True
                    ),

                "Charge":
                    st.column_config.NumberColumn(
                        "Charge",
                        format="$%.2f",
                        disabled=True
                    ),

                "Profit":
                    st.column_config.NumberColumn(
                        "Profit",
                        format="$%.2f",
                        disabled=True
                    )
            },

            key="resource_editor"
        )


        # Reactive calculations

        edited_resources = calculate_resources(
            edited_resources
        )

        task["resources"] = (
            edited_resources
        )

        task["description"] = (
            description
        )

        task["quantity"] = (
            quantity
        )

        task["unit"] = (
            unit
        )

        task["wbs"] = (
            wbs
        )


        task = calculate_task(
            task
        )

        st.session_state.active_task = task


        # ----------------------------------------------------
        # KPI BANNER
        # ----------------------------------------------------

        st.markdown(
            '<div class="section-header">'
            'ACTIVE TASK SUMMARY'
            '</div>',
            unsafe_allow_html=True
        )

        k = st.columns(8)

        k[0].metric(
            "Total Hours",
            f'{task["total_hours"]:,.2f}'
        )

        k[1].metric(
            "Labor",
            f'${task["labor_cost"]:,.2f}'
        )

        k[2].metric(
            "Materials",
            f'${task["material_cost"]:,.2f}'
        )

        k[3].metric(
            "Equipment",
            f'${task["equipment_cost"]:,.2f}'
        )

        k[4].metric(
            "Trucking",
            f'${task["hauling_cost"]:,.2f}'
        )

        k[5].metric(
            "Total Cost",
            f'${task["direct_cost"]:,.2f}'
        )

        k[6].metric(
            "Total Charge",
            f'${task["total_charge"]:,.2f}'
        )

        k[7].metric(
            "Profit",
            f'${task["profit"]:,.2f}'
        )


        # ----------------------------------------------------
        # BNI PRODUCTIVITY DERIVATION
        # ----------------------------------------------------

        if task.get(
            "bni_productivity",
            0
        ) > 0:

            bni_mh = float(
                task["bni_productivity"]
            )

            bni_total_hours = (
                quantity *
                bni_mh
            )

            st.markdown(
                '<div class="section-header">'
                'BNI PRODUCTIVITY DERIVATION'
                '</div>',
                unsafe_allow_html=True
            )

            st.code(
                f"Quantity × BNI Productivity\n"
                f"= {quantity:,.2f} {unit} × "
                f"{bni_mh:.4f} MH/{unit}\n\n"
                f"= {bni_total_hours:,.2f} "
                f"Total Man-Hours\n\n"
                f"BNI Reference Page: "
                f'{task.get("bni_page", "")}'
            )


        # ----------------------------------------------------
        # SAVE TASK
        # ----------------------------------------------------

        if st.button(
            "💾 SAVE / UPDATE ESTIMATE ITEM",
            type="primary"
        ):

            found = False

            for i, existing in enumerate(
                st.session_state.tasks
            ):

                if existing.get("id") == task.get(
                    "id"
                ):

                    st.session_state.tasks[i] = task

                    found = True

                    break


            if not found:

                st.session_state.tasks.append(
                    task
                )

            st.success(
                "Estimate item saved."
            )


# ============================================================
# SUMMARY / REPORTS
# ============================================================

with ribbon_tabs[3]:

    st.subheader(
        "📊 Summary / Reports"
    )

    if not st.session_state.tasks:

        st.info(
            "No estimate items have been saved."
        )

    else:

        summary_rows = []

        for i, task in enumerate(
            st.session_state.tasks,
            start=1
        ):

            summary_rows.append({

                "Line":
                    i,

                "WBS":
                    task.get(
                        "wbs",
                        ""
                    ),

                "Description":
                    task.get(
                        "description",
                        ""
                    ),

                "Qty":
                    task.get(
                        "quantity",
                        0
                    ),

                "Unit":
                    task.get(
                        "unit",
                        ""
                    ),

                "Hours":
                    task.get(
                        "total_hours",
                        0
                    ),

                "Labor":
                    task.get(
                        "labor_cost",
                        0
                    ),

                "Materials":
                    task.get(
                        "material_cost",
                        0
                    ),

                "Equipment":
                    task.get(
                        "equipment_cost",
                        0
                    ),

                "Trucking":
                    task.get(
                        "hauling_cost",
                        0
                    ),

                "Cost":
                    task.get(
                        "direct_cost",
                        0
                    ),

                "Charge":
                    task.get(
                        "total_charge",
                        0
                    ),

                "Profit":
                    task.get(
                        "profit",
                        0
                    )
            })


        summary_df = pd.DataFrame(
            summary_rows
        )

        st.dataframe(
            summary_df,
            use_container_width=True,
            hide_index=True
        )


        total_cost = summary_df[
            "Cost"
        ].sum()

        total_charge = summary_df[
            "Charge"
        ].sum()

        total_profit = summary_df[
            "Profit"
        ].sum()

        c1, c2, c3 = st.columns(3)

        c1.metric(
            "PROJECT COST",
            f"${total_cost:,.2f}"
        )

        c2.metric(
            "PROJECT CHARGE",
            f"${total_charge:,.2f}"
        )

        c3.metric(
            "PROJECT PROFIT",
            f"${total_profit:,.2f}"
        )


# ============================================================
# EXCEL EXPORT
# ============================================================

def export_excel(tasks):

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        # ----------------------------------------------------
        # PROJECT SUMMARY
        # ----------------------------------------------------

        summary_rows = []

        for i, task in enumerate(
            tasks,
            start=1
        ):

            summary_rows.append({

                "Line":
                    i,

                "WBS":
                    task.get("wbs", ""),

                "Description":
                    task.get(
                        "description",
                        ""
                    ),

                "Quantity":
                    task.get(
                        "quantity",
                        0
                    ),

                "Unit":
                    task.get(
                        "unit",
                        ""
                    ),

                "Total Hours":
                    task.get(
                        "total_hours",
                        0
                    ),

                "Labor":
                    task.get(
                        "labor_cost",
                        0
                    ),

                "Materials":
                    task.get(
                        "material_cost",
                        0
                    ),

                "Equipment":
                    task.get(
                        "equipment_cost",
                        0
                    ),

                "Trucking":
                    task.get(
                        "hauling_cost",
                        0
                    ),

                "Direct Cost":
                    task.get(
                        "direct_cost",
                        0
                    ),

                "Total Charge":
                    task.get(
                        "total_charge",
                        0
                    ),

                "Profit":
                    task.get(
                        "profit",
                        0
                    )
            })


        summary_df = pd.DataFrame(
            summary_rows
        )

        summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Estimate Summary"
        )


        # ----------------------------------------------------
        # RESOURCE BREAKDOWN
        # ----------------------------------------------------

        resource_rows = []

        for i, task in enumerate(
            tasks,
            start=1
        ):

            resources = task.get(
                "resources",
                pd.DataFrame()
            )

            if resources.empty:

                continue

            for _, row in resources.iterrows():

                resource_rows.append({

                    "Line":
                        i,

                    "WBS":
                        task.get(
                            "wbs",
                            ""
                        ),

                    "Task":
                        task.get(
                            "description",
                            ""
                        ),

                    "Type":
                        row.get(
                            "Type",
                            ""
                        ),

                    "Resource":
                        row.get(
                            "Resource",
                            ""
                        ),

                    "Quantity":
                        row.get(
                            "Quantity",
                            0
                        ),

                    "Unit":
                        row.get(
                            "Unit",
                            ""
                        ),

                    "Rate":
                        row.get(
                            "Rate",
                            0
                        ),

                    "Tax %":
                        row.get(
                            "Tax %",
                            0
                        ),

                    "Cost":
                        row.get(
                            "Cost",
                            0
                        ),

                    "Charge":
                        row.get(
                            "Charge",
                            0
                        ),

                    "Profit":
                        row.get(
                            "Profit",
                            0
                        )
                })


        resources_df = pd.DataFrame(
            resource_rows
        )

        resources_df.to_excel(
            writer,
            index=False,
            sheet_name="Resource Breakdown"
        )


        # ----------------------------------------------------
        # BNI PRODUCTIVITY
        # ----------------------------------------------------

        productivity_rows = []

        for i, task in enumerate(
            tasks,
            start=1
        ):

            quantity = float(
                task.get(
                    "quantity",
                    0
                )
            )

            productivity = float(
                task.get(
                    "bni_productivity",
                    0
                )
            )

            productivity_rows.append({

                "Line":
                    i,

                "Task":
                    task.get(
                        "description",
                        ""
                    ),

                "BNI Description":
                    task.get(
                        "bni_description",
                        ""
                    ),

                "BNI Page":
                    task.get(
                        "bni_page",
                        ""
                    ),

                "Quantity":
                    quantity,

                "Unit":
                    task.get(
                        "unit",
                        ""
                    ),

                "BNI MH / Unit":
                    productivity,

                "Total BNI MH":
                    quantity *
                    productivity
            })


        productivity_df = pd.DataFrame(
            productivity_rows
        )

        productivity_df.to_excel(
            writer,
            index=False,
            sheet_name="BNI Productivity"
        )


        # ----------------------------------------------------
        # MARKUP / TAX CONFIGURATION
        # ----------------------------------------------------

        settings_df = pd.DataFrame([

            [
                "Markup Multiplier",
                st.session_state.markup
            ],

            [
                "Markup %",
                (
                    st.session_state.markup
                    - 1
                )
                * 100
            ],

            [
                "Material Tax %",
                st.session_state.material_tax
                * 100
            ],

            [
                "Hauling Tax %",
                st.session_state.hauling_tax
                * 100
            ]

        ], columns=[
            "Setting",
            "Value"
        ])

        settings_df.to_excel(
            writer,
            index=False,
            sheet_name="Estimate Settings"
        )


        # ----------------------------------------------------
        # FORMATTING
        # ----------------------------------------------------

        workbook = writer.book

        from openpyxl.styles import (
            Font,
            PatternFill,
            Border,
            Side,
            Alignment
        )

        header_fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        section_fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

        total_fill = PatternFill(
            "solid",
            fgColor="E2F0D9"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        bold_font = Font(
            bold=True
        )

        thin = Side(
            style="thin",
            color="B7B7B7"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )


        for ws in workbook.worksheets:

            ws.freeze_panes = "A2"

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center"
                )
                cell.border = border


            for row in ws.iter_rows():

                for cell in row:

                    cell.border = border


            for column_cells in ws.columns:

                max_length = 0

                column_letter = (
                    column_cells[0]
                    .column_letter
                )

                for cell in column_cells:

                    try:

                        max_length = max(
                            max_length,
                            len(
                                str(
                                    cell.value
                                )
                            )
                        )

                    except Exception:

                        pass

                ws.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 2,
                    45
                )


        # ----------------------------------------------------
        # SPECIAL SUMMARY FORMATTING
        # ----------------------------------------------------

        if "Estimate Summary" in workbook.sheetnames:

            ws = workbook[
                "Estimate Summary"
            ]

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font


            last_row = ws.max_row

            for cell in ws[last_row]:

                cell.fill = total_fill
                cell.font = bold_font


    output.seek(0)

    return output.getvalue()


# ============================================================
# EXCEL EXPORT TAB
# ============================================================

with ribbon_tabs[4]:

    st.subheader(
        "📤 Excel Export"
    )

    if not st.session_state.tasks:

        st.info(
            "Save at least one estimate item first."
        )

    else:

        st.write(
            "The workbook contains:"
        )

        st.markdown("""
        - **Estimate Summary**
        - **Resource Breakdown**
        - **BNI Productivity**
        - **Estimate Settings**
        """)

        excel_data = export_excel(
            st.session_state.tasks
        )

        st.download_button(
            "📊 EXPORT PROFESSIONAL ESTIMATE — XLSX",
            data=excel_data,
            file_name=(
                "Civil_Estimate_"
                + datetime.now().strftime(
                    "%Y%m%d_%H%M"
                )
                + ".xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
            use_container_width=True
        )


# ============================================================
# FOOTER
# ============================================================

st.markdown(
    f"""
    <div class="status-bar">
        {APP_VERSION}
        &nbsp; | &nbsp;
        Markup: {st.session_state.markup:.2f}x
        &nbsp; | &nbsp;
        Material Tax: {st.session_state.material_tax:.1%}
        &nbsp; | &nbsp;
        Hauling Tax: {st.session_state.hauling_tax:.1%}
    </div>
    """,
    unsafe_allow_html=True
)
