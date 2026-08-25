import streamlit as st
import pandas as pd
from pathlib import Path
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Civil Estimating Software", page_icon="🏗️", layout="wide")
DB_DEFAULT = Path(__file__).parent / "construction_costs_pages_16_to_36.xlsx"
MARKUP_DEFAULT, MATERIAL_TAX_DEFAULT, HAULING_TAX_DEFAULT = 1.30, 0.09, 0.07
COLS = ["Type","Resource","Quantity","Unit","Rate","Tax %","Cost","Charge","Profit","Notes"]

st.markdown("""<style>
#MainMenu,footer,header{visibility:hidden}
.block-container{padding-top:.6rem;max-width:1500px}
.appbar{background:#1f4e78;color:white;padding:14px 18px;border-radius:6px;margin-bottom:8px}
.appbar-title{font-size:24px;font-weight:700}.appbar-sub{font-size:12px}
.step{background:#eaf2f8;border:1px solid #c8d9e6;border-radius:5px;padding:8px 11px;margin:8px 0;font-weight:650}
.help{background:#fff8e1;border:1px solid #f0d98c;border-radius:6px;padding:11px}
.total{background:#e2f0d9;border:1px solid #a9d18e;border-radius:6px;padding:12px}
</style>""", unsafe_allow_html=True)

def blank_resources():
    return pd.DataFrame(columns=COLS)

def init():
    defaults = {
        "tasks": [], "active_id": None,
        "project_name": "My Civil Construction Project", "estimator": "",
        "bni": None, "markup": MARKUP_DEFAULT,
        "material_tax": MATERIAL_TAX_DEFAULT, "hauling_tax": HAULING_TAX_DEFAULT,
        "labor": pd.DataFrame([["Foreman","HR",0.0],["Laborer","HR",0.0],["Equipment Operator","HR",0.0]],columns=["Resource","Unit","Rate"]),
        "materials": pd.DataFrame([['8" PVC',"LF",0.0],['8" Sewer Tap',"EA",0.0],["Crushed Gravel","CY",0.0],["Import Fill","CY",0.0]],columns=["Resource","Unit","Rate"]),
        "equipment": pd.DataFrame([["307.5","HR",0.0],["CTL","HR",0.0],["Excavator","HR",0.0],["Loader","HR",0.0]],columns=["Resource","Unit","Rate"]),
        "trucking": pd.DataFrame([["Tri-Axle","HR",0.0],["Dump Truck","HR",0.0]],columns=["Resource","Unit","Rate"]),
    }
    for k,v in defaults.items():
        if k not in st.session_state: st.session_state[k]=v
init()

def new_task():
    return {"id":datetime.now().timestamp(),"wbs":"33 — Utilities","description":"New Estimate Item",
            "quantity":1.0,"unit":"EA","surface":"Unpaved","depth":"",
            "bni_productivity":0.0,"bni_page":"","bni_description":"","resources":blank_resources()}

def active():
    for t in st.session_state.tasks:
        if t["id"]==st.session_state.active_id: return t
    return None

def save(t):
    for i,x in enumerate(st.session_state.tasks):
        if x["id"]==t["id"]: st.session_state.tasks[i]=t; return
    st.session_state.tasks.append(t)

def load_bni(src):
    df=pd.read_excel(src); df.columns=[str(c).strip() for c in df.columns]
    req=["CSI Code","Item Code","Description","Unit","Manhr/Unit","Page"]
    miss=[c for c in req if c not in df.columns]
    if miss: raise ValueError("Missing columns: "+", ".join(miss))
    df=df[req].copy()
    for c in ["CSI Code","Item Code","Description","Unit"]: df[c]=df[c].fillna("").astype(str).str.strip()
    df["Manhr/Unit"]=pd.to_numeric(df["Manhr/Unit"],errors="coerce")
    return df

def lookup(typ,name):
    tables={"Labor":st.session_state.labor,"Material":st.session_state.materials,
            "Equipment":st.session_state.equipment,"Trucking":st.session_state.trucking}
    df=tables.get(typ)
    if df is None or df.empty:return None
    m=df[df.Resource.astype(str).str.strip().str.lower()==name.strip().lower()]
    return None if m.empty else m.iloc[0]

def calc(df):
    if df is None or df.empty:return blank_resources()
    df=df.copy()
    for c in ["Type","Resource","Unit","Notes"]:
        if c not in df.columns:df[c]=""
        df[c]=df[c].fillna("").astype(str)
    for c in ["Quantity","Rate","Tax %"]:
        if c not in df.columns:df[c]=0.0
        df[c]=pd.to_numeric(df[c],errors="coerce").fillna(0.0)
    costs=[];charges=[];profits=[]
    for _,r in df.iterrows():
        cost=float(r.Quantity)*float(r.Rate)*(1+float(r["Tax %"]))
        charge=cost*st.session_state.markup
        costs.append(cost);charges.append(charge);profits.append(charge-cost)
    df["Cost"]=costs;df["Charge"]=charges;df["Profit"]=profits
    return df[COLS]

def task_calc(t):
    t=dict(t);t["resources"]=calc(t["resources"]);r=t["resources"]
    def sub(k): return r.loc[r.Type.str.lower()==k.lower(),"Cost"].sum() if not r.empty else 0
    t["labor_cost"]=sub("Labor");t["material_cost"]=sub("Material");t["equipment_cost"]=sub("Equipment");t["trucking_cost"]=sub("Trucking")
    t["direct_cost"]=t["labor_cost"]+t["material_cost"]+t["equipment_cost"]+t["trucking_cost"]
    t["total_charge"]=r["Charge"].sum() if not r.empty else 0;t["profit"]=t["total_charge"]-t["direct_cost"]
    t["bni_hours"]=float(t["quantity"])*float(t["bni_productivity"])
    return t

def export_xlsx():
    out=BytesIO()
    with pd.ExcelWriter(out,engine="openpyxl") as w:
        summary=[];resources=[];bni=[]
        for i,raw in enumerate(st.session_state.tasks,1):
            t=task_calc(raw)
            summary.append({"Line":i,"WBS":t["wbs"],"Description":t["description"],"Quantity":t["quantity"],"Unit":t["unit"],
                            "Surface":t["surface"],"Depth":t["depth"],"BNI MH/Unit":t["bni_productivity"],
                            "BNI Page":t["bni_page"],"BNI Total MH":t["bni_hours"],"Labor":t["labor_cost"],
                            "Materials":t["material_cost"],"Equipment":t["equipment_cost"],"Trucking":t["trucking_cost"],
                            "Direct Cost":t["direct_cost"],"Total Charge":t["total_charge"],"Profit":t["profit"]})
            for _,r in t["resources"].iterrows():
                resources.append({"Line":i,"Task":t["description"],"Type":r.Type,"Resource":r.Resource,"Quantity":r.Quantity,
                                   "Unit":r.Unit,"Rate":r.Rate,"Tax %":r["Tax %"],"Cost":r.Cost,"Charge":r.Charge,"Profit":r.Profit,"Notes":r.Notes})
            bni.append({"Line":i,"Task":t["description"],"BNI Description":t["bni_description"],"BNI Page":t["bni_page"],
                        "Quantity":t["quantity"],"Unit":t["unit"],"BNI MH/Unit":t["bni_productivity"],
                        "Total BNI Man-Hours":t["bni_hours"],
                        "Derivation":f'{t["quantity"]:,.2f} {t["unit"]} × {t["bni_productivity"]:.4f} MH/{t["unit"]}'})
        pd.DataFrame(summary).to_excel(w,index=False,sheet_name="Estimate Summary")
        pd.DataFrame(resources).to_excel(w,index=False,sheet_name="Resource Breakdown")
        pd.DataFrame(bni).to_excel(w,index=False,sheet_name="BNI Productivity")
        pd.DataFrame([["Project",st.session_state.project_name],["Estimator",st.session_state.estimator],
                      ["Markup",st.session_state.markup],["Material Tax",st.session_state.material_tax],
                      ["Hauling Tax",st.session_state.hauling_tax]],columns=["Setting","Value"]).to_excel(w,index=False,sheet_name="Settings")
        for name,df in [("Labor Price List",st.session_state.labor),("Material Price List",st.session_state.materials),
                        ("Equipment Price List",st.session_state.equipment),("Trucking Price List",st.session_state.trucking)]:
            df.to_excel(w,index=False,sheet_name=name)
        wb=w.book;fill=PatternFill("solid",fgColor="1F4E78");font=Font(color="FFFFFF",bold=True);thin=Side(style="thin",color="B7B7B7")
        for ws in wb.worksheets:
            ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]:
                cell.fill=fill;cell.font=font;cell.alignment=Alignment(horizontal="center")
            for col in ws.columns:
                letter=get_column_letter(col[0].column);ws.column_dimensions[letter].width=min(max(len(str(x.value or "")) for x in col)+2,45)
    out.seek(0);return out.getvalue()

st.markdown('<div class="appbar"><div class="appbar-title">🏗️ Civil Estimating Software</div><div class="appbar-sub">Guided Estimating Workbook • BNI Productivity • Resource Cost Analysis</div></div>',unsafe_allow_html=True)
tabs=st.tabs(["🏠 Home","📋 Estimate","📚 Libraries","📊 Reports","📤 Export"])

with tabs[0]:
    st.header("Welcome 👋")
    st.markdown('<div class="help"><b>Simple workflow:</b> Start → Define scope → Select BNI productivity → Add resources → Review cost → Export Excel.</div>',unsafe_allow_html=True)
    if st.button("➕ START NEW ESTIMATE",type="primary",use_container_width=True):
        t=new_task();st.session_state.tasks.append(t);st.session_state.active_id=t["id"];st.rerun()
    st.subheader("Current Project")
    st.session_state.project_name=st.text_input("Project Name",st.session_state.project_name)
    st.session_state.estimator=st.text_input("Estimator",st.session_state.estimator)
    st.write(f"Estimate items: **{len(st.session_state.tasks)}**")
    if st.session_state.tasks:
        rows=[]
        for i,x in enumerate(st.session_state.tasks,1):
            t=task_calc(x);rows.append({"Line":i,"Description":t["description"],"Qty":t["quantity"],"Unit":t["unit"],"Cost":t["direct_cost"],"Charge":t["total_charge"],"Profit":t["profit"]})
        st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

with tabs[1]:
    st.header("Estimate Builder")
    if not st.session_state.tasks:
        st.info("No estimate items yet. Go to Home and click START NEW ESTIMATE.")
    else:
        st.sidebar.markdown("### 📂 Estimate Items")
        for i,t in enumerate(st.session_state.tasks,1):
            if st.sidebar.button(f"{i}. {t['description']}"[:42],key=f"side_{t['id']}",use_container_width=True):
                st.session_state.active_id=t["id"];st.rerun()
        t=active()
        if t:
            st.markdown('<div class="step">STEP 1 — DEFINE THE SCOPE</div>',unsafe_allow_html=True)
            t["description"]=st.text_input("Scope of Work",t["description"],help='Example: Installation of 2,050 LF of 8" PVC Sanitary Line')
            c1,c2,c3=st.columns(3)
            with c1:t["quantity"]=st.number_input("Quantity",min_value=0.0,value=float(t["quantity"]),step=1.0)
            with c2:
                units=["LF","SF","CY","TON","EA","HR","LS"];t["unit"]=st.selectbox("Unit",units,index=units.index(t["unit"]) if t["unit"] in units else 0)
            with c3:
                surfaces=["Paved","Unpaved","Repaired"];t["surface"]=st.selectbox("Surface",surfaces,index=surfaces.index(t["surface"]) if t["surface"] in surfaces else 0)
            c1,c2=st.columns(2)
            with c1:t["wbs"]=st.selectbox("MasterFormat / WBS",["02 — Existing Conditions","03 — Concrete","31 — Earthwork","32 — Exterior Improvements","33 — Utilities"],index=4 if t["wbs"]=="33 — Utilities" else 0)
            with c2:t["depth"]=st.text_input("Average Depth",t["depth"],placeholder="Example: 0'-6'")

            st.markdown('<div class="step">STEP 2 — SELECT BNI PRODUCTIVITY</div>',unsafe_allow_html=True)
            if st.session_state.bni is None:
                st.warning("Go to 📚 Libraries → BNI Productivity and upload your BNI Excel database.")
            else:
                b=st.session_state.bni;q=st.text_input("Search BNI",placeholder="sanitary / PVC / sewer / excavation")
                s=b.copy()
                if q.strip():
                    ql=q.lower().strip()
                    s=s[s.Description.str.lower().str.contains(ql,na=False)|s["CSI Code"].str.lower().str.contains(ql,na=False)|s["Item Code"].str.lower().str.contains(ql,na=False)]
                s=s[s["Manhr/Unit"].notna()].head(150)
                if not s.empty:
                    opts={idx:f'{r.Description} | {r.Unit} | {r["Manhr/Unit"]:.4f} MH/{r.Unit} | Page {r.Page}' for idx,r in s.iterrows()}
                    sel=st.selectbox("Choose BNI item",list(opts),format_func=lambda x:opts[x])
                    br=b.loc[sel];t["bni_productivity"]=float(br["Manhr/Unit"]);t["bni_page"]=str(br.Page);t["bni_description"]=str(br.Description)
                    st.success(f'BNI productivity: {t["bni_productivity"]:.4f} MH/{t["unit"]} • Page {t["bni_page"]} • Total BNI hours: {t["quantity"]*t["bni_productivity"]:,.2f}')
                    st.code(f'{t["quantity"]:,.2f} {t["unit"]} × {t["bni_productivity"]:.4f} MH/{t["unit"]} = {t["quantity"]*t["bni_productivity"]:,.2f} BNI man-hours')
                else:st.info("No matching BNI item found.")

            st.markdown('<div class="step">STEP 3 — ADD LABOR, EQUIPMENT, MATERIALS & TRUCKING</div>',unsafe_allow_html=True)
            st.caption("Click + Add Row. Choose a type, type a resource name, and enter quantity/rate. Known library items automatically receive their unit, rate and tax.")
            ed=st.data_editor(t["resources"],num_rows="dynamic",use_container_width=True,hide_index=True,key=f"editor_{t['id']}",
                column_config={"Type":st.column_config.SelectboxColumn("Type",options=["Labor","Equipment","Material","Trucking"]),
                               "Quantity":st.column_config.NumberColumn("Qty",min_value=0.0),
                               "Rate":st.column_config.NumberColumn("Rate",format="$%.2f"),
                               "Tax %":st.column_config.NumberColumn("Tax %",format="%.1f%%"),
                               "Cost":st.column_config.NumberColumn("Cost",format="$%.2f",disabled=True),
                               "Charge":st.column_config.NumberColumn("Charge",format="$%.2f",disabled=True),
                               "Profit":st.column_config.NumberColumn("Profit",format="$%.2f",disabled=True)})
            ed=ed.copy()
            for idx in ed.index:
                typ=str(ed.at[idx,"Type"]).strip();name=str(ed.at[idx,"Resource"]).strip()
                if name:
                    found=lookup(typ,name)
                    if found is not None:
                        if not str(ed.at[idx,"Unit"]).strip():ed.at[idx,"Unit"]=found.Unit
                        if float(ed.at[idx,"Rate"] or 0)==0:ed.at[idx,"Rate"]=float(found.Rate)
                    if float(ed.at[idx,"Tax %"] or 0)==0:
                        ed.at[idx,"Tax %"]=st.session_state.material_tax if typ=="Material" else st.session_state.hauling_tax if typ=="Trucking" else 0.0
            t["resources"]=calc(ed);t=task_calc(t)

            st.markdown('<div class="step">STEP 4 — REVIEW</div>',unsafe_allow_html=True)
            cs=st.columns(7)
            vals=[("BNI Hours",t["bni_hours"],False),("Labor",t["labor_cost"],True),("Materials",t["material_cost"],True),("Equipment",t["equipment_cost"],True),("Trucking",t["trucking_cost"],True),("Total Cost",t["direct_cost"],True),("Profit",t["profit"],True)]
            for col,(lab,val,money) in zip(cs,vals):col.metric(lab,f"${val:,.2f}" if money else f"{val:,.2f}")
            st.markdown(f'<div class="total"><b>TOTAL CHARGE / BID VALUE</b><div style="font-size:28px;font-weight:700">${t["total_charge"]:,.2f}</div></div>',unsafe_allow_html=True)

            c1,c2,c3=st.columns(3)
            with c1:
                if st.button("💾 SAVE ITEM",type="primary",use_container_width=True):save(t);st.success("Estimate item saved.")
            with c2:
                if st.button("📋 DUPLICATE",use_container_width=True):
                    cp=dict(t);cp["id"]=datetime.now().timestamp();cp["description"]+=" — Copy";cp["resources"]=t["resources"].copy();st.session_state.tasks.append(cp);st.session_state.active_id=cp["id"];st.rerun()
            with c3:
                if st.button("🗑️ DELETE",use_container_width=True):
                    st.session_state.tasks=[x for x in st.session_state.tasks if x["id"]!=t["id"]];st.session_state.active_id=st.session_state.tasks[0]["id"] if st.session_state.tasks else None;st.rerun()

with tabs[2]:
    st.header("📚 Libraries")
    lt=st.tabs(["BNI Productivity","Labor","Materials","Equipment","Trucking","Settings"])
    with lt[0]:
        up=st.file_uploader("Upload BNI Excel database",type=["xlsx"],help="Required columns: CSI Code, Item Code, Description, Unit, Manhr/Unit, Page.")
        if up:
            try:st.session_state.bni=load_bni(up);st.success(f"Loaded {len(st.session_state.bni):,} BNI records.")
            except Exception as e:st.error(str(e))
        if st.session_state.bni is None and DB_DEFAULT.exists():
            try:st.session_state.bni=load_bni(DB_DEFAULT)
            except Exception:pass
        if st.session_state.bni is not None:st.dataframe(st.session_state.bni.head(200),use_container_width=True,hide_index=True)
    with lt[1]:st.session_state.labor=st.data_editor(st.session_state.labor,num_rows="dynamic",use_container_width=True,hide_index=True,key="lib_labor")
    with lt[2]:st.session_state.materials=st.data_editor(st.session_state.materials,num_rows="dynamic",use_container_width=True,hide_index=True,key="lib_material")
    with lt[3]:st.session_state.equipment=st.data_editor(st.session_state.equipment,num_rows="dynamic",use_container_width=True,hide_index=True,key="lib_equipment")
    with lt[4]:st.session_state.trucking=st.data_editor(st.session_state.trucking,num_rows="dynamic",use_container_width=True,hide_index=True,key="lib_trucking")
    with lt[5]:
        st.session_state.markup=st.number_input("Markup Multiplier",min_value=1.0,value=float(st.session_state.markup),step=.01)
        st.session_state.material_tax=st.number_input("Material Tax",min_value=0.,max_value=1.,value=float(st.session_state.material_tax),step=.01,format="%.2f")
        st.session_state.hauling_tax=st.number_input("Hauling Tax",min_value=0.,max_value=1.,value=float(st.session_state.hauling_tax),step=.01,format="%.2f")

with tabs[3]:
    st.header("📊 Reports")
    if not st.session_state.tasks:st.info("No saved estimate items.")
    else:
        rows=[]
        for i,x in enumerate(st.session_state.tasks,1):
            t=task_calc(x);rows.append({"Line":i,"WBS":t["wbs"],"Description":t["description"],"Qty":t["quantity"],"Unit":t["unit"],"BNI Hours":t["bni_hours"],"Labor":t["labor_cost"],"Materials":t["material_cost"],"Equipment":t["equipment_cost"],"Trucking":t["trucking_cost"],"Direct Cost":t["direct_cost"],"Charge":t["total_charge"],"Profit":t["profit"]})
        df=pd.DataFrame(rows);st.dataframe(df,use_container_width=True,hide_index=True)
        c=st.columns(4);c[0].metric("BNI Hours",f'{df["BNI Hours"].sum():,.2f}');c[1].metric("Direct Cost",f'${df["Direct Cost"].sum():,.2f}');c[2].metric("Charge",f'${df["Charge"].sum():,.2f}');c[3].metric("Profit",f'${df["Profit"].sum():,.2f}')

with tabs[4]:
    st.header("📤 Export")
    if not st.session_state.tasks:st.info("Create and save an estimate item first.")
    else:
        st.write("Excel includes Estimate Summary, Resource Breakdown, BNI Productivity, Settings and price-list sheets.")
        st.download_button("📊 DOWNLOAD EXCEL ESTIMATE",export_xlsx(),file_name=f'Civil_Estimate_{datetime.now():%Y%m%d_%H%M}.xlsx',mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary",use_container_width=True)

with st.sidebar:
    st.markdown("### 🏗️ Quick Start")
    if st.button("➕ New Estimate Item",use_container_width=True):
        t=new_task();st.session_state.tasks.append(t);st.session_state.active_id=t["id"];st.rerun()
    st.divider()
    st.markdown("**Estimate Items**")
    if not st.session_state.tasks:st.caption("None yet.")
    for i,t in enumerate(st.session_state.tasks,1):
        if st.button(f"{i}. {t['description']}"[:40],key=f"quick_{t['id']}",use_container_width=True):
            st.session_state.active_id=t["id"];st.rerun()
    st.divider()
    st.caption("Version 8.1 • Guided workflow")
